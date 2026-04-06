from __future__ import annotations

from typing import Iterable

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from ._models import DEFAULT_FILL_VALUE, CellWrite, FillPlan
from ._ranges import contains_cell, is_excluded, parse_range, parse_ranges


def build_fill_plan(
    worksheet: Worksheet,
    *,
    target_range: str,
    excluded_ranges: Iterable[str] | None = None,
    fill_value: str = DEFAULT_FILL_VALUE,
    merge_empty_runs: bool = False,
    preserved_coordinates: set[tuple[int, int]] | None = None,
) -> FillPlan:
    target = parse_range(target_range)
    exclusions = parse_ranges(excluded_ranges)
    fill_text = str(fill_value)
    merge_lookup = build_merge_lookup(worksheet)
    protected_coordinates = preserved_coordinates or set()

    existing_merge_writes, filled_from_existing_merges = plan_existing_merged_anchor_writes(
        worksheet=worksheet,
        target=target,
        exclusions=exclusions,
        fill_text=fill_text,
        preserved_coordinates=protected_coordinates,
    )
    plain_writes, created_merges, filled_from_plain_cells = plan_plain_cell_writes(
        worksheet=worksheet,
        target=target,
        exclusions=exclusions,
        merge_lookup=merge_lookup,
        fill_text=fill_text,
        merge_empty_runs=merge_empty_runs,
        preserved_coordinates=protected_coordinates,
    )

    return FillPlan(
        sheet_name=worksheet.title,
        target_range=target.coord,
        fill_value=fill_text,
        filled_cells=filled_from_existing_merges + filled_from_plain_cells,
        merged_ranges=tuple(created_merges),
        cell_writes=tuple(existing_merge_writes + plain_writes),
    )


def apply_fill_plan_to_worksheet(worksheet: Worksheet, plan: FillPlan) -> None:
    for merged_range in plan.merged_ranges:
        worksheet.merge_cells(merged_range)

    for cell_write in plan.cell_writes:
        worksheet.cell(row=cell_write.row, column=cell_write.column).value = cell_write.value


def resolve_worksheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name is None:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Worksheet {sheet_name!r} was not found. Available sheets: {available}")
    return workbook[sheet_name]


def plan_existing_merged_anchor_writes(
    worksheet: Worksheet,
    *,
    target: CellRange,
    exclusions: tuple[CellRange, ...],
    fill_text: str,
    preserved_coordinates: set[tuple[int, int]],
) -> tuple[list[CellWrite], int]:
    cell_writes: list[CellWrite] = []
    filled_cells = 0

    for merged_range in worksheet.merged_cells.ranges:
        row = merged_range.min_row
        column = merged_range.min_col
        if not contains_cell(target, row, column):
            continue
        if is_excluded(exclusions, row, column):
            continue
        if (row, column) in preserved_coordinates:
            continue

        cell = worksheet.cell(row=row, column=column)
        if is_empty(cell.value) and not has_preserved_comment(cell):
            cell_writes.append(CellWrite(row=row, column=column, value=fill_text))
            filled_cells += 1

    return cell_writes, filled_cells


def plan_plain_cell_writes(
    worksheet: Worksheet,
    *,
    target: CellRange,
    exclusions: tuple[CellRange, ...],
    merge_lookup: dict[tuple[int, int], tuple[int, int]],
    fill_text: str,
    merge_empty_runs: bool,
    preserved_coordinates: set[tuple[int, int]],
) -> tuple[list[CellWrite], list[str], int]:
    cell_writes: list[CellWrite] = []
    created_merges: list[str] = []
    filled_cells = 0

    for column in range(target.min_col, target.max_col + 1):
        empty_run: list[tuple[int, int]] = []
        for row in range(target.min_row, target.max_row + 1):
            coordinate = (row, column)
            if is_fillable_plain_cell(
                worksheet=worksheet,
                row=row,
                column=column,
                exclusions=exclusions,
                merge_lookup=merge_lookup,
                preserved_coordinates=preserved_coordinates,
            ):
                empty_run.append(coordinate)
                continue

            flushed_cells, flushed_writes, merged_range = flush_empty_run(
                empty_run=empty_run,
                fill_text=fill_text,
                merge_empty_runs=merge_empty_runs,
            )
            filled_cells += flushed_cells
            cell_writes.extend(flushed_writes)
            if merged_range is not None:
                created_merges.append(merged_range)
            empty_run = []

        flushed_cells, flushed_writes, merged_range = flush_empty_run(
            empty_run=empty_run,
            fill_text=fill_text,
            merge_empty_runs=merge_empty_runs,
        )
        filled_cells += flushed_cells
        cell_writes.extend(flushed_writes)
        if merged_range is not None:
            created_merges.append(merged_range)

    return cell_writes, created_merges, filled_cells


def is_fillable_plain_cell(
    worksheet: Worksheet,
    *,
    row: int,
    column: int,
    exclusions: tuple[CellRange, ...],
    merge_lookup: dict[tuple[int, int], tuple[int, int]],
    preserved_coordinates: set[tuple[int, int]],
) -> bool:
    if is_excluded(exclusions, row, column):
        return False
    if (row, column) in merge_lookup:
        return False
    if (row, column) in preserved_coordinates:
        return False
    cell = worksheet.cell(row=row, column=column)
    return is_empty(cell.value) and not has_preserved_comment(cell)


def flush_empty_run(
    *,
    empty_run: list[tuple[int, int]],
    fill_text: str,
    merge_empty_runs: bool,
) -> tuple[int, list[CellWrite], str | None]:
    if not empty_run:
        return 0, [], None

    if merge_empty_runs and len(empty_run) >= 2:
        start_row, column = empty_run[0]
        end_row, _ = empty_run[-1]
        range_string = f"{get_column_letter(column)}{start_row}:{get_column_letter(column)}{end_row}"
        return len(empty_run), [CellWrite(row=start_row, column=column, value=fill_text)], range_string

    return (
        len(empty_run),
        [CellWrite(row=row, column=column, value=fill_text) for row, column in empty_run],
        None,
    )


def build_merge_lookup(worksheet: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, column)] = anchor
    return lookup


def is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def has_preserved_comment(cell: object) -> bool:
    return getattr(cell, "comment", None) is not None
