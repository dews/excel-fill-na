from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_fill_na.cli import main
from conftest import FIXTURE_PATH, artifact_payloads, worksheet_cell_xml


def test_cli_applies_fill_merge_and_output_settings(tmp_path: Path, capsys) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A4"] = "stop"
    workbook.save(source)

    exit_code = main(
        [
            str(source),
            "--sheet",
            "Data",
            "--range",
            "A1:A4",
            "--merge-empty-runs",
            "--output",
            str(output),
            "--fill-text",
            "MISSING",
        ]
    )

    assert exit_code == 0

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]
    merged_ranges = {str(cell_range) for cell_range in output_worksheet.merged_cells.ranges}
    captured = capsys.readouterr()

    assert "A1:A3" in merged_ranges
    assert output_worksheet["A1"].value == "MISSING"
    assert "created 1 merged range" in captured.out


def test_cli_preserves_fixture_artifacts(tmp_path: Path, capsys) -> None:
    output = tmp_path / "fixture-output.xlsx"
    source_workbook = load_workbook(FIXTURE_PATH)
    source_worksheet = source_workbook["Data"]
    source_image_count = len(getattr(source_worksheet, "_images", []))
    source_chart_count = len(getattr(source_worksheet, "_charts", []))
    source_workbook.close()

    exit_code = main(
        [
            str(FIXTURE_PATH),
            "--sheet",
            "Data",
            "--range",
            "B1:C1",
            "--output",
            str(output),
            "--fill-text",
            "CLI",
        ]
    )

    assert exit_code == 0

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]
    captured = capsys.readouterr()

    assert output_worksheet["B1"].value == "CLI"
    assert output_worksheet["C1"].value == "CLI"
    assert len(getattr(output_worksheet, "_images", [])) == source_image_count
    assert len(getattr(output_worksheet, "_charts", [])) == source_chart_count
    assert artifact_payloads(output) == artifact_payloads(FIXTURE_PATH)
    cell_attributes, cell_value = worksheet_cell_xml(output, "D3")
    assert cell_attributes["t"] == "e"
    assert cell_attributes["vm"] == "1"
    assert cell_value == "#VALUE!"
    assert "filled 2 empty cells" in captured.out


def test_cli_deletes_empty_rows(tmp_path: Path, capsys) -> None:
    source = tmp_path / "delete-cli.xlsx"
    output = tmp_path / "delete-cli.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A2"] = "keep"
    workbook.save(source)
    workbook.close()

    exit_code = main(
        [
            str(source),
            "--sheet",
            "Data",
            "--range",
            "A1:A2",
            "--delete-empty-rows",
            "--output",
            str(output),
        ]
    )

    assert exit_code == 0

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]
    captured = capsys.readouterr()

    assert output_worksheet["A1"].value == "keep"
    assert "deleted 1 empty row" in captured.out
    output_workbook.close()


def test_cli_rejects_delete_mode_with_merge_mode(capsys) -> None:
    with pytest.raises(SystemExit) as excinfo:
        main(["input.xlsx", "--range", "A1:A2", "--delete-empty-rows", "--merge-empty-runs"])

    captured = capsys.readouterr()

    assert excinfo.value.code == 2
    assert "--delete-empty-rows cannot be combined with --merge-empty-runs" in captured.err


def test_cli_rejects_delete_mode_with_explicit_fill_text(capsys) -> None:
    with pytest.raises(SystemExit) as excinfo:
        main(["input.xlsx", "--range", "A1:A2", "--delete-empty-rows", "--fill-text", "NA"])

    captured = capsys.readouterr()

    assert excinfo.value.code == 2
    assert "--delete-empty-rows cannot be combined with --fill-text" in captured.err
