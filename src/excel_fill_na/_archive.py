from __future__ import annotations

from bisect import bisect_left, bisect_right
from io import BytesIO
from pathlib import Path
import posixpath
import re
from shutil import copyfile
from tempfile import NamedTemporaryFile
from typing import Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.cell_range import CellRange

from ._models import FillPlan
from ._ranges import parse_range

SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DRAWINGML_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
VML_NS = "urn:schemas-microsoft-com:vml"
EXCEL_VML_NS = "urn:schemas-microsoft-com:office:excel"
THREADED_COMMENTS_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
THREADED_COMMENT_REL_NS = "http://schemas.microsoft.com/office/2017/10/relationships/threadedComment"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

COMMENTS_REL_TYPE = f"{OFFICE_DOCUMENT_REL_NS}/comments"
DRAWING_REL_TYPE = f"{OFFICE_DOCUMENT_REL_NS}/drawing"
VML_DRAWING_REL_TYPE = f"{OFFICE_DOCUMENT_REL_NS}/vmlDrawing"

SHEET_NAMESPACES = {"main": SPREADSHEETML_NS}
WORKBOOK_NAMESPACES = {
    "main": SPREADSHEETML_NS,
    "r": OFFICE_DOCUMENT_REL_NS,
    "rels": PACKAGE_REL_NS,
}
DRAWING_NAMESPACES = {"xdr": DRAWINGML_NS}


def persist_workbook_changes(
    *,
    source: Path,
    destination: Path,
    worksheet_path: str,
    plan: FillPlan,
) -> None:
    same_path = source.resolve() == destination.resolve()
    has_changes = bool(plan.cell_writes or plan.merged_ranges or plan.deleted_row_indices)
    if not has_changes:
        if same_path:
            return
        copyfile(source, destination)
        return

    if same_path:
        write_patched_archive_in_place(
            source=source,
            destination=destination,
            worksheet_path=worksheet_path,
            plan=plan,
        )
        return

    write_patched_archive(
        source=source,
        destination=destination,
        worksheet_path=worksheet_path,
        plan=plan,
    )


def resolve_worksheet_archive_path(source: Path, sheet_name: str) -> str:
    with ZipFile(source) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    relationship_targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships_root.findall(package_relationship_tag("Relationship"))
    }

    sheets = workbook_root.find("main:sheets", WORKBOOK_NAMESPACES)
    if sheets is None:
        raise ValueError("Workbook metadata is missing sheet definitions.")

    relationship_attr = f"{{{OFFICE_DOCUMENT_REL_NS}}}id"
    for sheet in sheets.findall("main:sheet", WORKBOOK_NAMESPACES):
        if sheet.attrib.get("name") != sheet_name:
            continue

        relationship_id = sheet.attrib.get(relationship_attr)
        if relationship_id is None or relationship_id not in relationship_targets:
            raise ValueError(f"Worksheet {sheet_name!r} is missing workbook relationship metadata.")
        return normalize_archive_path("xl/workbook.xml", relationship_targets[relationship_id])

    raise ValueError(f"Worksheet {sheet_name!r} was not found in workbook metadata.")


def find_value_metadata_cells(source: Path, worksheet_path: str) -> set[tuple[int, int]]:
    with ZipFile(source) as archive:
        worksheet_root = ET.fromstring(archive.read(worksheet_path))

    sheet_data = worksheet_root.find("main:sheetData", SHEET_NAMESPACES)
    if sheet_data is None:
        return set()

    coordinates: set[tuple[int, int]] = set()
    for row_element in sheet_data.findall(sheet_tag("row")):
        for cell_element in row_element.findall(sheet_tag("c")):
            coordinate = cell_element.attrib.get("r")
            if coordinate is None or "vm" not in cell_element.attrib:
                continue
            coordinates.add(
                (
                    coordinate_row_index(coordinate),
                    coordinate_column_index(coordinate),
                )
            )

    return coordinates


def write_patched_archive_in_place(
    *,
    source: Path,
    destination: Path,
    worksheet_path: str,
    plan: FillPlan,
) -> None:
    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            prefix=f"{source.stem}.",
            suffix=source.suffix,
            dir=source.parent,
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
        write_patched_archive(
            source=source,
            destination=temporary_path,
            worksheet_path=worksheet_path,
            plan=plan,
        )
        temporary_path.replace(destination)
    except Exception:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()
        raise


def write_patched_archive(
    *,
    source: Path,
    destination: Path,
    worksheet_path: str,
    plan: FillPlan,
) -> None:
    with ZipFile(source) as source_archive:
        patched_parts = build_archive_patches(
            source_archive=source_archive,
            worksheet_path=worksheet_path,
            plan=plan,
        )

        with ZipFile(destination, "w") as destination_archive:
            for info in source_archive.infolist():
                data = patched_parts.get(info.filename, source_archive.read(info.filename))
                destination_archive.writestr(info, data)


def build_archive_patches(
    *,
    source_archive: ZipFile,
    worksheet_path: str,
    plan: FillPlan,
) -> dict[str, bytes]:
    patched_parts = {
        worksheet_path: patch_worksheet_xml(source_archive.read(worksheet_path), plan),
    }

    if not plan.deleted_row_indices:
        return patched_parts

    related_part_paths = resolve_related_part_paths(source_archive, worksheet_path)
    for path in related_part_paths.get(COMMENTS_REL_TYPE, ()):
        patched_parts[path] = patch_comments_xml(source_archive.read(path), plan.deleted_row_indices)
    for path in related_part_paths.get(THREADED_COMMENT_REL_NS, ()):
        patched_parts[path] = patch_threaded_comments_xml(source_archive.read(path), plan.deleted_row_indices)
    for path in related_part_paths.get(VML_DRAWING_REL_TYPE, ()):
        patched_parts[path] = patch_vml_drawing_xml(source_archive.read(path), plan.deleted_row_indices)
    for path in related_part_paths.get(DRAWING_REL_TYPE, ()):
        patched_parts[path] = patch_drawing_xml(source_archive.read(path), plan.deleted_row_indices)

    return patched_parts


def resolve_related_part_paths(source_archive: ZipFile, worksheet_path: str) -> dict[str, tuple[str, ...]]:
    relationships_path = relationship_part_path(worksheet_path)
    try:
        relationships_root = ET.fromstring(source_archive.read(relationships_path))
    except KeyError:
        return {}

    paths_by_type: dict[str, list[str]] = {}
    for relationship in relationships_root.findall(package_relationship_tag("Relationship")):
        relationship_type = relationship.attrib.get("Type")
        target = relationship.attrib.get("Target")
        if not relationship_type or not target:
            continue
        paths_by_type.setdefault(relationship_type, []).append(
            normalize_archive_path(worksheet_path, target)
        )

    return {
        relationship_type: tuple(paths)
        for relationship_type, paths in paths_by_type.items()
    }


def patch_worksheet_xml(worksheet_xml: bytes, plan: FillPlan) -> bytes:
    worksheet_root, namespaces = parse_xml_bytes(worksheet_xml)
    register_namespaces(namespaces)

    sheet_data = worksheet_root.find("main:sheetData", SHEET_NAMESPACES)
    if sheet_data is None:
        raise ValueError("Worksheet XML is missing sheetData.")

    if plan.deleted_row_indices:
        delete_rows_from_sheet_data(sheet_data, plan.deleted_row_indices)
        patch_existing_merge_ranges(worksheet_root, plan.deleted_row_indices)
        patch_worksheet_references(worksheet_root, plan.deleted_row_indices)

    row_lookup = {
        safe_int(row.attrib.get("r")): row
        for row in sheet_data.findall(sheet_tag("row"))
        if safe_int(row.attrib.get("r")) is not None
    }
    touched_rows: set[int] = set()

    for cell_write in plan.cell_writes:
        row_element = get_or_create_row(sheet_data, row_lookup, cell_write.row)
        cell_element = get_or_create_cell(row_element, cell_write.row, cell_write.column)
        set_inline_string_value(cell_element, cell_write.value)
        touched_rows.add(cell_write.row)

    for row_index in touched_rows:
        update_row_spans(row_lookup[row_index])

    if plan.merged_ranges:
        append_merge_ranges(worksheet_root, plan.merged_ranges)

    update_dimension(worksheet_root, sheet_data, plan)

    return serialize_xml(
        worksheet_root,
        namespaces,
        xml_declaration=has_xml_declaration(worksheet_xml),
    )


def patch_comments_xml(comments_xml: bytes, deleted_row_indices: tuple[int, ...]) -> bytes:
    comments_root, namespaces = parse_xml_bytes(comments_xml)
    register_namespaces(namespaces)

    comment_list = comments_root.find("main:commentList", SHEET_NAMESPACES)
    if comment_list is not None:
        for comment in list(comment_list.findall(sheet_tag("comment"))):
            reference = comment.attrib.get("ref")
            if not reference:
                continue
            shifted_reference = shift_single_coordinate(reference, deleted_row_indices)
            if shifted_reference is None:
                comment_list.remove(comment)
                continue
            comment.set("ref", shifted_reference)

    return serialize_xml(
        comments_root,
        namespaces,
        xml_declaration=has_xml_declaration(comments_xml),
    )


def patch_threaded_comments_xml(
    threaded_comments_xml: bytes,
    deleted_row_indices: tuple[int, ...],
) -> bytes:
    threaded_root, namespaces = parse_xml_bytes(threaded_comments_xml)
    register_namespaces(namespaces)

    for threaded_comment in list(
        threaded_root.findall(f"{{{THREADED_COMMENTS_NS}}}threadedComment")
    ):
        reference = threaded_comment.attrib.get("ref")
        if not reference:
            continue
        shifted_reference = shift_single_coordinate(reference, deleted_row_indices)
        if shifted_reference is None:
            threaded_root.remove(threaded_comment)
            continue
        threaded_comment.set("ref", shifted_reference)

    return serialize_xml(
        threaded_root,
        namespaces,
        xml_declaration=has_xml_declaration(threaded_comments_xml),
    )


def patch_vml_drawing_xml(vml_xml: bytes, deleted_row_indices: tuple[int, ...]) -> bytes:
    vml_root, namespaces = parse_xml_bytes(vml_xml)
    register_namespaces(namespaces)

    for shape in list(vml_root):
        if shape.tag != vml_tag("shape"):
            continue

        client_data = shape.find(excel_vml_tag("ClientData"))
        if client_data is None or client_data.attrib.get("ObjectType") != "Note":
            continue

        row_element = client_data.find(excel_vml_tag("Row"))
        if row_element is None or row_element.text is None:
            continue

        row_index = safe_int(row_element.text)
        if row_index is None:
            continue

        shifted_row = shift_zero_based_row_marker(row_index, deleted_row_indices)
        if shifted_row is None:
            vml_root.remove(shape)
            continue

        row_element.text = str(shifted_row)

        anchor = client_data.find(excel_vml_tag("Anchor"))
        if anchor is not None and anchor.text:
            anchor.text = shift_vml_anchor(anchor.text, deleted_row_indices)

    return serialize_xml(
        vml_root,
        namespaces,
        xml_declaration=has_xml_declaration(vml_xml),
    )


def patch_drawing_xml(drawing_xml: bytes, deleted_row_indices: tuple[int, ...]) -> bytes:
    drawing_root, namespaces = parse_xml_bytes(drawing_xml)
    register_namespaces(namespaces)

    for one_cell_anchor in drawing_root.findall("xdr:oneCellAnchor", DRAWING_NAMESPACES):
        shift_drawing_marker(one_cell_anchor.find("xdr:from", DRAWING_NAMESPACES), deleted_row_indices)

    for two_cell_anchor in drawing_root.findall("xdr:twoCellAnchor", DRAWING_NAMESPACES):
        from_marker = two_cell_anchor.find("xdr:from", DRAWING_NAMESPACES)
        to_marker = two_cell_anchor.find("xdr:to", DRAWING_NAMESPACES)
        from_row = shift_drawing_marker(from_marker, deleted_row_indices)
        to_row = shift_drawing_marker(to_marker, deleted_row_indices)
        if from_row is not None and to_row is not None and to_row < from_row:
            row_element = to_marker.find("xdr:row", DRAWING_NAMESPACES) if to_marker is not None else None
            if row_element is not None:
                row_element.text = str(from_row)

    return serialize_xml(
        drawing_root,
        namespaces,
        xml_declaration=has_xml_declaration(drawing_xml),
    )


def delete_rows_from_sheet_data(sheet_data: ET.Element, deleted_row_indices: tuple[int, ...]) -> None:
    for row_element in list(sheet_data.findall(sheet_tag("row"))):
        row_index = safe_int(row_element.attrib.get("r"))
        if row_index is None:
            continue

        shifted_row = shift_row_number(row_index, deleted_row_indices)
        if shifted_row is None:
            sheet_data.remove(row_element)
            continue

        row_element.set("r", str(shifted_row))
        for cell_element in row_element.findall(sheet_tag("c")):
            coordinate = cell_element.attrib.get("r")
            if coordinate is None:
                continue
            shifted_coordinate = shift_single_coordinate(coordinate, deleted_row_indices)
            if shifted_coordinate is not None:
                cell_element.set("r", shifted_coordinate)
        update_row_spans(row_element)


def patch_existing_merge_ranges(worksheet_root: ET.Element, deleted_row_indices: tuple[int, ...]) -> None:
    merge_cells = worksheet_root.find("main:mergeCells", SHEET_NAMESPACES)
    if merge_cells is None:
        return

    for merge_cell in list(merge_cells.findall(sheet_tag("mergeCell"))):
        reference = merge_cell.attrib.get("ref")
        if not reference:
            continue
        shifted_bounds = shift_range_bounds(reference, deleted_row_indices)
        if shifted_bounds is None or range_is_single_cell(shifted_bounds):
            merge_cells.remove(merge_cell)
            continue
        merge_cell.set("ref", format_range_bounds(shifted_bounds))

    if any(child.tag == sheet_tag("mergeCell") for child in merge_cells):
        merge_cells.set(
            "count",
            str(sum(1 for child in merge_cells if child.tag == sheet_tag("mergeCell"))),
        )
        return

    worksheet_root.remove(merge_cells)


def patch_worksheet_references(worksheet_root: ET.Element, deleted_row_indices: tuple[int, ...]) -> None:
    patch_hyperlinks(worksheet_root, deleted_row_indices)
    patch_selection_references(worksheet_root, deleted_row_indices)


def patch_hyperlinks(worksheet_root: ET.Element, deleted_row_indices: tuple[int, ...]) -> None:
    hyperlinks = worksheet_root.find("main:hyperlinks", SHEET_NAMESPACES)
    if hyperlinks is None:
        return

    for hyperlink in list(hyperlinks.findall(sheet_tag("hyperlink"))):
        reference = hyperlink.attrib.get("ref")
        if not reference:
            continue
        shifted_bounds = shift_range_bounds(reference, deleted_row_indices)
        if shifted_bounds is None:
            hyperlinks.remove(hyperlink)
            continue
        hyperlink.set("ref", format_range_bounds(shifted_bounds))

    if not any(child.tag == sheet_tag("hyperlink") for child in hyperlinks):
        worksheet_root.remove(hyperlinks)


def patch_selection_references(worksheet_root: ET.Element, deleted_row_indices: tuple[int, ...]) -> None:
    for selection in worksheet_root.findall(".//main:selection", SHEET_NAMESPACES):
        active_cell = selection.attrib.get("activeCell")
        if active_cell:
            shifted_active_cell = shift_single_coordinate(active_cell, deleted_row_indices)
            if shifted_active_cell is None:
                selection.attrib.pop("activeCell", None)
            else:
                selection.set("activeCell", shifted_active_cell)

        sqref = selection.attrib.get("sqref")
        if sqref:
            shifted_sqref = shift_reference_list(sqref, deleted_row_indices)
            if shifted_sqref is None:
                selection.attrib.pop("sqref", None)
            else:
                selection.set("sqref", shifted_sqref)


def shift_reference_list(reference_list: str, deleted_row_indices: tuple[int, ...]) -> str | None:
    shifted_references = [
        format_range_bounds(shifted_bounds)
        for reference in reference_list.split()
        if (shifted_bounds := shift_range_bounds(reference, deleted_row_indices)) is not None
    ]
    if not shifted_references:
        return None
    return " ".join(shifted_references)


def normalize_archive_path(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_path), target))


def relationship_part_path(owner_path: str) -> str:
    return posixpath.join(
        posixpath.dirname(owner_path),
        "_rels",
        posixpath.basename(owner_path) + ".rels",
    )


def parse_xml_bytes(xml_bytes: bytes) -> tuple[ET.Element, list[tuple[str, str]]]:
    namespaces: list[tuple[str, str]] = []
    seen_namespaces: set[tuple[str, str]] = set()
    for _, namespace in ET.iterparse(BytesIO(xml_bytes), events=("start-ns",)):
        if namespace not in seen_namespaces:
            namespaces.append(namespace)
            seen_namespaces.add(namespace)
    return ET.fromstring(xml_bytes), namespaces


def register_namespaces(namespaces: Iterable[tuple[str, str]]) -> None:
    for prefix, uri in namespaces:
        if re.match(r"ns\d+$", prefix):
            continue
        ET.register_namespace(prefix, uri)


def serialize_xml(
    root: ET.Element,
    namespaces: Iterable[tuple[str, str]],
    *,
    xml_declaration: bool,
) -> bytes:
    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=xml_declaration)
    return restore_root_namespace_declarations(xml_bytes, namespaces)


def has_xml_declaration(xml_bytes: bytes) -> bool:
    return xml_bytes.lstrip().startswith(b"<?xml")


def restore_root_namespace_declarations(
    xml_bytes: bytes,
    namespaces: Iterable[tuple[str, str]],
) -> bytes:
    xml_text = xml_bytes.decode("utf-8")
    root_match = re.search(r"<([A-Za-z_][^>\s/]*)\b[^>]*>", xml_text)
    if root_match is None:
        return xml_bytes

    opening_tag = root_match.group(0)
    declarations: list[str] = []
    for prefix, uri in namespaces:
        declaration = f'xmlns:{prefix}="{uri}"' if prefix else f'xmlns="{uri}"'
        if declaration not in opening_tag:
            declarations.append(declaration)

    if not declarations:
        return xml_bytes

    insertion = " " + " ".join(declarations)
    updated_text = xml_text.replace(opening_tag, opening_tag[:-1] + insertion + ">", 1)
    return updated_text.encode("utf-8")


def get_or_create_row(
    sheet_data: ET.Element,
    row_lookup: dict[int, ET.Element],
    row_index: int,
) -> ET.Element:
    existing_row = row_lookup.get(row_index)
    if existing_row is not None:
        return existing_row

    row_element = ET.Element(sheet_tag("row"), {"r": str(row_index)})
    insert_at = len(sheet_data)
    for position, candidate in enumerate(sheet_data):
        candidate_row = safe_int(candidate.attrib.get("r"))
        if candidate_row is not None and candidate_row > row_index:
            insert_at = position
            break
    sheet_data.insert(insert_at, row_element)
    row_lookup[row_index] = row_element
    return row_element


def get_or_create_cell(row_element: ET.Element, row_index: int, column_index: int) -> ET.Element:
    coordinate = f"{get_column_letter(column_index)}{row_index}"
    insert_at = len(row_element)

    for position, candidate in enumerate(row_element):
        if candidate.tag != sheet_tag("c"):
            continue

        candidate_coordinate = candidate.attrib.get("r")
        if candidate_coordinate == coordinate:
            return candidate
        if candidate_coordinate is not None and coordinate_column_index(candidate_coordinate) > column_index:
            insert_at = position
            break

    cell_element = ET.Element(sheet_tag("c"), {"r": coordinate})
    row_element.insert(insert_at, cell_element)
    return cell_element


def set_inline_string_value(cell_element: ET.Element, value: str) -> None:
    preserved_attributes = {
        name: attribute_value
        for name, attribute_value in cell_element.attrib.items()
        if name not in ("t", "vm")
    }

    for child in list(cell_element):
        cell_element.remove(child)

    cell_element.attrib.clear()
    cell_element.attrib.update(preserved_attributes)
    cell_element.set("t", "inlineStr")

    inline_string = ET.SubElement(cell_element, sheet_tag("is"))
    text_element = ET.SubElement(inline_string, sheet_tag("t"))
    if value != value.strip():
        text_element.set(XML_SPACE, "preserve")
    text_element.text = value


def append_merge_ranges(worksheet_root: ET.Element, merged_ranges: Iterable[str]) -> None:
    merge_cells = worksheet_root.find("main:mergeCells", SHEET_NAMESPACES)
    if merge_cells is None:
        merge_cells = ET.Element(sheet_tag("mergeCells"))
        insert_merge_cells_element(worksheet_root, merge_cells)

    existing_ranges = {
        merge_cell.attrib.get("ref")
        for merge_cell in merge_cells.findall(sheet_tag("mergeCell"))
    }
    for merged_range in merged_ranges:
        if merged_range not in existing_ranges:
            ET.SubElement(merge_cells, sheet_tag("mergeCell"), {"ref": merged_range})
            existing_ranges.add(merged_range)

    merge_cells.set(
        "count",
        str(sum(1 for child in merge_cells if child.tag == sheet_tag("mergeCell"))),
    )


def insert_merge_cells_element(worksheet_root: ET.Element, merge_cells: ET.Element) -> None:
    preceding_names = {
        "sheetData",
        "sheetCalcPr",
        "sheetProtection",
        "protectedRanges",
        "scenarios",
        "autoFilter",
        "sortState",
        "dataConsolidate",
        "customSheetViews",
    }
    insert_at: int | None = None
    for index, child in enumerate(list(worksheet_root)):
        if local_name(child.tag) in preceding_names:
            insert_at = index + 1

    if insert_at is None:
        worksheet_root.append(merge_cells)
        return

    worksheet_root.insert(insert_at, merge_cells)


def update_dimension(worksheet_root: ET.Element, sheet_data: ET.Element, plan: FillPlan) -> None:
    if plan.deleted_row_indices:
        bounds = scan_sheet_data_bounds(sheet_data)
        bounds = expand_bounds_with_existing_merge_ranges(bounds, worksheet_root)
    else:
        bounds = existing_dimension_bounds(worksheet_root)
        if bounds is None:
            bounds = scan_sheet_data_bounds(sheet_data)

    for cell_write in plan.cell_writes:
        bounds = expand_bounds(bounds, row=cell_write.row, column=cell_write.column)
    for merged_range in plan.merged_ranges:
        bounds = expand_bounds_with_range(bounds, parse_range(merged_range))

    ref = "A1" if bounds is None else format_dimension(bounds)
    dimension = worksheet_root.find("main:dimension", SHEET_NAMESPACES)
    if dimension is None:
        dimension = ET.Element(sheet_tag("dimension"), {"ref": ref})
        insert_dimension_element(worksheet_root, dimension)
        return

    dimension.set("ref", ref)


def existing_dimension_bounds(worksheet_root: ET.Element) -> tuple[int, int, int, int] | None:
    dimension = worksheet_root.find("main:dimension", SHEET_NAMESPACES)
    if dimension is None:
        return None

    ref = dimension.attrib.get("ref")
    if not ref:
        return None

    try:
        cell_range = parse_range(ref)
    except ValueError:
        return None

    return (
        cell_range.min_row,
        cell_range.min_col,
        cell_range.max_row,
        cell_range.max_col,
    )


def scan_sheet_data_bounds(sheet_data: ET.Element) -> tuple[int, int, int, int] | None:
    bounds: tuple[int, int, int, int] | None = None
    for row_element in sheet_data.findall(sheet_tag("row")):
        row_index = safe_int(row_element.attrib.get("r"))
        if row_index is None:
            continue

        for cell_element in row_element.findall(sheet_tag("c")):
            coordinate = cell_element.attrib.get("r")
            if coordinate is None:
                continue
            bounds = expand_bounds(
                bounds,
                row=row_index,
                column=coordinate_column_index(coordinate),
            )
    return bounds


def expand_bounds_with_existing_merge_ranges(
    bounds: tuple[int, int, int, int] | None,
    worksheet_root: ET.Element,
) -> tuple[int, int, int, int] | None:
    merge_cells = worksheet_root.find("main:mergeCells", SHEET_NAMESPACES)
    if merge_cells is None:
        return bounds

    for merge_cell in merge_cells.findall(sheet_tag("mergeCell")):
        reference = merge_cell.attrib.get("ref")
        if not reference:
            continue
        bounds = expand_bounds_with_range(bounds, parse_range(reference))
    return bounds


def insert_dimension_element(worksheet_root: ET.Element, dimension: ET.Element) -> None:
    insert_at = 0
    for index, child in enumerate(list(worksheet_root)):
        if local_name(child.tag) == "sheetPr":
            insert_at = index + 1
            break
        insert_at = index
        break
    worksheet_root.insert(insert_at, dimension)


def expand_bounds(
    bounds: tuple[int, int, int, int] | None,
    *,
    row: int,
    column: int,
) -> tuple[int, int, int, int]:
    if bounds is None:
        return (row, column, row, column)

    min_row, min_col, max_row, max_col = bounds
    return (
        min(min_row, row),
        min(min_col, column),
        max(max_row, row),
        max(max_col, column),
    )


def expand_bounds_with_range(
    bounds: tuple[int, int, int, int] | None,
    cell_range: CellRange,
) -> tuple[int, int, int, int]:
    bounds = expand_bounds(bounds, row=cell_range.min_row, column=cell_range.min_col)
    return expand_bounds(bounds, row=cell_range.max_row, column=cell_range.max_col)


def format_dimension(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def update_row_spans(row_element: ET.Element) -> None:
    columns = [
        coordinate_column_index(cell.attrib["r"])
        for cell in row_element
        if cell.tag == sheet_tag("c") and "r" in cell.attrib
    ]
    if not columns:
        row_element.attrib.pop("spans", None)
        return

    row_element.set("spans", f"{min(columns)}:{max(columns)}")


def shift_range_bounds(
    reference: str,
    deleted_row_indices: tuple[int, ...],
) -> tuple[int, int, int, int] | None:
    try:
        cell_range = parse_range(reference)
    except ValueError:
        return None

    first_surviving_row: int | None = None
    last_surviving_row: int | None = None
    for row_index in range(cell_range.min_row, cell_range.max_row + 1):
        if shift_row_number(row_index, deleted_row_indices) is None:
            continue
        first_surviving_row = row_index
        break

    if first_surviving_row is None:
        return None

    for row_index in range(cell_range.max_row, cell_range.min_row - 1, -1):
        if shift_row_number(row_index, deleted_row_indices) is None:
            continue
        last_surviving_row = row_index
        break

    assert last_surviving_row is not None
    shifted_min_row = shift_row_number(first_surviving_row, deleted_row_indices)
    shifted_max_row = shift_row_number(last_surviving_row, deleted_row_indices)
    assert shifted_min_row is not None
    assert shifted_max_row is not None

    return (
        shifted_min_row,
        cell_range.min_col,
        shifted_max_row,
        cell_range.max_col,
    )


def format_range_bounds(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def range_is_single_cell(bounds: tuple[int, int, int, int]) -> bool:
    min_row, min_col, max_row, max_col = bounds
    return min_row == max_row and min_col == max_col


def shift_single_coordinate(coordinate: str, deleted_row_indices: tuple[int, ...]) -> str | None:
    column_letters, row_index = coordinate_from_string(coordinate)
    shifted_row = shift_row_number(row_index, deleted_row_indices)
    if shifted_row is None:
        return None
    return f"{column_letters}{shifted_row}"


def shift_row_number(row_index: int, deleted_row_indices: tuple[int, ...]) -> int | None:
    insertion_point = bisect_left(deleted_row_indices, row_index)
    if insertion_point < len(deleted_row_indices) and deleted_row_indices[insertion_point] == row_index:
        return None
    return row_index - insertion_point


def shift_zero_based_row_marker(
    row_marker: int,
    deleted_row_indices: tuple[int, ...],
) -> int | None:
    row_index = row_marker + 1
    deleted_before_or_at = bisect_right(deleted_row_indices, row_index)
    shifted_row = row_index - deleted_before_or_at
    if shifted_row < 1:
        return None
    return shifted_row - 1


def shift_vml_anchor(anchor_text: str, deleted_row_indices: tuple[int, ...]) -> str:
    values = [part.strip() for part in anchor_text.split(",")]
    for row_position in (2, 6):
        if row_position >= len(values):
            continue
        row_value = safe_int(values[row_position])
        if row_value is None:
            continue
        shifted_row = shift_zero_based_row_marker(row_value, deleted_row_indices)
        values[row_position] = str(0 if shifted_row is None else shifted_row)
    return ", ".join(values)


def shift_drawing_marker(
    marker: ET.Element | None,
    deleted_row_indices: tuple[int, ...],
) -> int | None:
    if marker is None:
        return None

    row_element = marker.find("xdr:row", DRAWING_NAMESPACES)
    if row_element is None or row_element.text is None:
        return None

    row_value = safe_int(row_element.text)
    if row_value is None:
        return None

    shifted_row = shift_zero_based_row_marker(row_value, deleted_row_indices)
    if shifted_row is None:
        shifted_row = 0
    row_element.text = str(shifted_row)
    return shifted_row


def coordinate_column_index(coordinate: str) -> int:
    column_letters, _ = coordinate_from_string(coordinate)
    return column_index_from_string(column_letters)


def coordinate_row_index(coordinate: str) -> int:
    _, row_index = coordinate_from_string(coordinate)
    return row_index


def sheet_tag(local_name: str) -> str:
    return f"{{{SPREADSHEETML_NS}}}{local_name}"


def package_relationship_tag(local_name: str) -> str:
    return f"{{{PACKAGE_REL_NS}}}{local_name}"


def vml_tag(local_name: str) -> str:
    return f"{{{VML_NS}}}{local_name}"


def excel_vml_tag(local_name: str) -> str:
    return f"{{{EXCEL_VML_NS}}}{local_name}"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def safe_int(value: str | None) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except ValueError:
        return None
