from __future__ import annotations

from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

DEFAULT_FILL_VALUE = "NA"


@dataclass(frozen=True, slots=True)
class FillResult:
    sheet_name: str
    target_range: str
    fill_value: str
    filled_cells: int
    merged_ranges: tuple[str, ...]
    output_path: Path | None = None


def process_workbook(
    input_path: str | Path,
    *,
    target_range: str,
    excluded_ranges: Iterable[str] | None = None,
    fill_value: str = DEFAULT_FILL_VALUE,
    merge_empty_runs: bool = False,
    sheet_name: str | None = None,
    output_path: str | Path | None = None,
) -> FillResult:
    """Load a workbook, fill empty cells, and save the result."""
    source = Path(input_path)
    if not source.exists():
        raise FileNotFoundError(f"Workbook not found: {source}")

    destination = (
        Path(output_path)
        if output_path is not None
        else source.with_name(f"{source.stem}.filled{source.suffix}")
    )

    workbook = load_workbook(
        filename=source,
        keep_vba=source.suffix.lower() == ".xlsm",
    )
    worksheet = _resolve_worksheet(workbook, sheet_name)
    result = fill_empty_cells(
        worksheet,
        target_range=target_range,
        excluded_ranges=excluded_ranges,
        fill_value=fill_value,
        merge_empty_runs=merge_empty_runs,
    )
    workbook.save(destination)
    return replace(result, output_path=destination)


def fill_empty_cells(
    worksheet: Worksheet,
    *,
    target_range: str,
    excluded_ranges: Iterable[str] | None = None,
    fill_value: str = DEFAULT_FILL_VALUE,
    merge_empty_runs: bool = False,
) -> FillResult:
    """Fill empty cells inside a worksheet range."""
    target = _parse_range(target_range)
    exclusions = _parse_ranges(excluded_ranges)
    fill_text = str(fill_value)
    merge_lookup = _build_merge_lookup(worksheet)

    filled_from_existing_merges = _fill_existing_merged_anchors(
        worksheet=worksheet,
        target=target,
        exclusions=exclusions,
        fill_text=fill_text,
    )
    filled_from_plain_cells, created_merges = _fill_plain_cells(
        worksheet=worksheet,
        target=target,
        exclusions=exclusions,
        merge_lookup=merge_lookup,
        fill_text=fill_text,
        merge_empty_runs=merge_empty_runs,
    )

    return FillResult(
        sheet_name=worksheet.title,
        target_range=target.coord,
        fill_value=fill_text,
        filled_cells=filled_from_existing_merges + filled_from_plain_cells,
        merged_ranges=tuple(created_merges),
    )


def _resolve_worksheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name is None:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Worksheet {sheet_name!r} was not found. Available sheets: {available}")
    return workbook[sheet_name]


def _parse_ranges(range_strings: Iterable[str] | None) -> tuple[CellRange, ...]:
    if not range_strings:
        return ()

    parsed_ranges: list[CellRange] = []
    for range_string in range_strings:
        for candidate in str(range_string).split(","):
            cleaned = candidate.strip()
            if cleaned:
                parsed_ranges.append(_parse_range(cleaned))
    return tuple(parsed_ranges)


def _parse_range(range_string: str) -> CellRange:
    try:
        return CellRange(range_string)
    except ValueError as exc:
        raise ValueError(f"Invalid cell range: {range_string!r}") from exc


def _fill_existing_merged_anchors(
    worksheet: Worksheet,
    *,
    target: CellRange,
    exclusions: tuple[CellRange, ...],
    fill_text: str,
) -> int:
    filled_cells = 0
    for merged_range in worksheet.merged_cells.ranges:
        row = merged_range.min_row
        col = merged_range.min_col
        if not _contains_cell(target, row, col):
            continue
        if _is_excluded(exclusions, row, col):
            continue

        cell = worksheet.cell(row=row, column=col)
        if _is_empty(cell.value):
            cell.value = fill_text
            filled_cells += 1
    return filled_cells


def _fill_plain_cells(
    worksheet: Worksheet,
    *,
    target: CellRange,
    exclusions: tuple[CellRange, ...],
    merge_lookup: dict[tuple[int, int], tuple[int, int]],
    fill_text: str,
    merge_empty_runs: bool,
) -> tuple[int, list[str]]:
    filled_cells = 0
    created_merges: list[str] = []

    for column in range(target.min_col, target.max_col + 1):
        empty_run: list[tuple[int, int]] = []
        for row in range(target.min_row, target.max_row + 1):
            coordinate = (row, column)
            if _is_fillable_plain_cell(
                worksheet=worksheet,
                row=row,
                column=column,
                exclusions=exclusions,
                merge_lookup=merge_lookup,
            ):
                empty_run.append(coordinate)
                continue

            filled_cells += _flush_empty_run(
                worksheet=worksheet,
                empty_run=empty_run,
                fill_text=fill_text,
                merge_empty_runs=merge_empty_runs,
                created_merges=created_merges,
            )
            empty_run = []

        filled_cells += _flush_empty_run(
            worksheet=worksheet,
            empty_run=empty_run,
            fill_text=fill_text,
            merge_empty_runs=merge_empty_runs,
            created_merges=created_merges,
        )

    return filled_cells, created_merges


def _is_fillable_plain_cell(
    worksheet: Worksheet,
    *,
    row: int,
    column: int,
    exclusions: tuple[CellRange, ...],
    merge_lookup: dict[tuple[int, int], tuple[int, int]],
) -> bool:
    if _is_excluded(exclusions, row, column):
        return False
    if (row, column) in merge_lookup:
        return False
    return _is_empty(worksheet.cell(row=row, column=column).value)


def _flush_empty_run(
    worksheet: Worksheet,
    *,
    empty_run: list[tuple[int, int]],
    fill_text: str,
    merge_empty_runs: bool,
    created_merges: list[str],
) -> int:
    if not empty_run:
        return 0

    if merge_empty_runs and len(empty_run) >= 2:
        start_row, column = empty_run[0]
        end_row, _ = empty_run[-1]
        range_string = f"{get_column_letter(column)}{start_row}:{get_column_letter(column)}{end_row}"
        worksheet.merge_cells(range_string)
        worksheet.cell(row=start_row, column=column).value = fill_text
        created_merges.append(range_string)
        return len(empty_run)

    for row, column in empty_run:
        worksheet.cell(row=row, column=column).value = fill_text
    return len(empty_run)


def _build_merge_lookup(worksheet: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, column)] = anchor
    return lookup


def _contains_cell(cell_range: CellRange, row: int, column: int) -> bool:
    return (
        cell_range.min_row <= row <= cell_range.max_row
        and cell_range.min_col <= column <= cell_range.max_col
    )


def _is_excluded(exclusions: tuple[CellRange, ...], row: int, column: int) -> bool:
    return any(_contains_cell(cell_range, row, column) for cell_range in exclusions)


def _is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")

