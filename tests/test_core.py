from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_from_string

from excel_fill_na._archive import find_value_metadata_cells, resolve_worksheet_archive_path
from excel_fill_na._planning import build_delete_plan
from excel_fill_na.core import fill_empty_cells, process_workbook
from conftest import (
    FIXTURE_PATH,
    artifact_payloads,
    comment_refs,
    drawing_anchor_rows,
    root_namespace_declarations,
    threaded_comment_refs,
    worksheet_cell_xml,
    worksheet_hyperlink_refs,
    worksheet_merge_refs,
)


def _write_comment_only_workbook(path: Path) -> None:
    base = path.with_name("comment-only.base.xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = None
    worksheet["A2"] = "placeholder"
    worksheet["A2"].comment = Comment("note", "tester")
    worksheet["A3"] = "stop"
    workbook.save(base)
    workbook.close()

    with ZipFile(base) as source_archive:
        sheet_root = ET.fromstring(source_archive.read("xl/worksheets/sheet1.xml"))
        comments_root = ET.fromstring(source_archive.read("xl/comments/comment1.xml"))

        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        row = sheet_root.find("main:sheetData/main:row[@r='2']", namespace)
        assert row is not None

        cell = row.find("main:c[@r='A2']", namespace)
        assert cell is not None
        for child in list(cell):
            cell.remove(child)
        cell.attrib.clear()
        cell.set("r", "A2")

        comment = comments_root.find("main:commentList/main:comment", namespace)
        assert comment is not None
        comment.set("ref", "A2")

        with ZipFile(path, "w", compression=ZIP_DEFLATED) as destination_archive:
            for info in source_archive.infolist():
                data = source_archive.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    data = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
                elif info.filename == "xl/comments/comment1.xml":
                    data = ET.tostring(comments_root, encoding="utf-8", xml_declaration=True)
                destination_archive.writestr(info, data)

    base.unlink()


def _write_value_metadata_only_workbook(path: Path) -> None:
    base = path.with_name("value-metadata-only.base.xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = None
    worksheet["A2"] = "placeholder"
    worksheet["A3"] = "stop"
    workbook.save(base)
    workbook.close()

    with ZipFile(base) as source_archive:
        sheet_root = ET.fromstring(source_archive.read("xl/worksheets/sheet1.xml"))

        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        row = sheet_root.find("main:sheetData/main:row[@r='2']", namespace)
        assert row is not None

        cell = row.find("main:c[@r='A2']", namespace)
        assert cell is not None
        for child in list(cell):
            cell.remove(child)
        cell.attrib.clear()
        cell.set("r", "A2")
        cell.set("vm", "1")

        with ZipFile(path, "w", compression=ZIP_DEFLATED) as destination_archive:
            for info in source_archive.infolist():
                data = source_archive.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    data = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
                destination_archive.writestr(info, data)

    base.unlink()


def _shift_coordinate_after_deleting_rows(
    coordinate: str,
    deleted_rows: tuple[int, ...],
) -> str:
    column_letters, row_index = coordinate_from_string(coordinate)
    shifted_row = row_index - sum(1 for deleted_row in deleted_rows if deleted_row < row_index)
    return f"{column_letters}{shifted_row}"


def _shift_zero_based_anchor_after_deleting_rows(
    row_index: int | None,
    deleted_rows: tuple[int, ...],
) -> int | None:
    if row_index is None:
        return None
    shifted_row = row_index - sum(1 for deleted_row in deleted_rows if deleted_row <= row_index + 1)
    return max(0, shifted_row)


def test_fill_empty_cells_only_in_selected_range() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A2"] = "keep"

    result = fill_empty_cells(worksheet, target_range="A1:A2")

    assert worksheet["A1"].value == "NA"
    assert worksheet["A2"].value == "keep"
    assert worksheet["B1"].value is None
    assert result.filled_cells == 1
    assert result.merged_ranges == ()


def test_fill_empty_cells_respects_excluded_ranges() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["B2"] = "keep"

    result = fill_empty_cells(
        worksheet,
        target_range="A1:B2",
        excluded_ranges=["A2", "B1:B2"],
        fill_value="MISSING",
    )

    assert worksheet["A1"].value == "MISSING"
    assert worksheet["A2"].value is None
    assert worksheet["B1"].value is None
    assert worksheet["B2"].value == "keep"
    assert result.filled_cells == 1


def test_merge_empty_runs_creates_vertical_merge() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A4"] = "stop"

    result = fill_empty_cells(
        worksheet,
        target_range="A1:A4",
        merge_empty_runs=True,
    )

    merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}

    assert "A1:A3" in merged_ranges
    assert worksheet["A1"].value == "NA"
    assert worksheet["A4"].value == "stop"
    assert result.filled_cells == 3
    assert result.merged_ranges == ("A1:A3",)


def test_existing_merged_cells_are_supported() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("B1:B3")

    result = fill_empty_cells(
        worksheet,
        target_range="B1:B3",
        fill_value="EMPTY",
    )

    assert worksheet["B1"].value == "EMPTY"
    assert result.filled_cells == 1
    assert result.merged_ranges == ()


def test_fill_empty_cells_keeps_comment_only_cells_unfilled_and_unmerged() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A2"].comment = Comment("note", "tester")
    worksheet["A3"] = "stop"

    result = fill_empty_cells(
        worksheet,
        target_range="A1:A3",
        merge_empty_runs=True,
    )

    assert result.filled_cells == 1
    assert result.merged_ranges == ()
    assert not worksheet.merged_cells.ranges
    assert worksheet["A1"].value == "NA"
    assert worksheet["A2"].value is None
    assert worksheet["A2"].comment is not None
    assert worksheet["A2"].comment.text == "note"
    assert worksheet["A3"].value == "stop"


def test_process_workbook_saves_default_output_file(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    workbook.active["C3"] = None
    workbook.save(source)

    result = process_workbook(
        source,
        target_range="C3",
        fill_value="EMPTY",
    )

    assert result.output_path == tmp_path / "input.filled.xlsx"
    output_workbook = load_workbook(result.output_path)
    assert output_workbook.active["C3"].value == "EMPTY"


def test_process_workbook_preserves_fixture_artifacts_and_styles(tmp_path: Path) -> None:
    output = tmp_path / "preserved.xlsx"

    source_workbook = load_workbook(FIXTURE_PATH)
    source_worksheet = source_workbook["Data"]
    source_style_id = source_worksheet["C1"].style_id
    source_image_count = len(getattr(source_worksheet, "_images", []))
    source_chart_count = len(getattr(source_worksheet, "_charts", []))
    source_workbook.close()

    result = process_workbook(
        FIXTURE_PATH,
        sheet_name="Data",
        target_range="B1:C1",
        fill_value="EMPTY",
        output_path=output,
    )

    assert result.filled_cells == 2

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert output_worksheet["B1"].value == "EMPTY"
    assert output_worksheet["C1"].value == "EMPTY"
    assert output_worksheet["C1"].style_id == source_style_id
    assert len(getattr(output_worksheet, "_images", [])) == source_image_count
    assert len(getattr(output_worksheet, "_charts", [])) == source_chart_count

    output_artifacts = artifact_payloads(output)
    assert "xl/charts/chart1.xml" in output_artifacts
    assert "xl/richData/richValueRel.xml" in output_artifacts
    assert "xl/worksheets/_rels/sheet1.xml.rels" in output_artifacts
    assert output_artifacts == artifact_payloads(FIXTURE_PATH)
    assert root_namespace_declarations(output) == root_namespace_declarations(FIXTURE_PATH)
    cell_attributes, cell_value = worksheet_cell_xml(output, "D3")
    assert cell_attributes["t"] == "e"
    assert cell_attributes["vm"] == "1"
    assert cell_value == "#VALUE!"


def test_process_workbook_targets_named_sheet_without_touching_other_sheets(tmp_path: Path) -> None:
    output = tmp_path / "other-sheet.xlsx"

    result = process_workbook(
        FIXTURE_PATH,
        sheet_name="Other",
        target_range="B2",
        fill_value="FILLED",
        output_path=output,
    )

    assert result.sheet_name == "Other"
    assert result.filled_cells == 1

    output_workbook = load_workbook(output)
    assert output_workbook["Other"]["A1"].value == "keep"
    assert output_workbook["Other"]["B2"].value == "FILLED"
    assert output_workbook["Data"]["B1"].value is None
    assert output_workbook["Data"]["C1"].value is None
    assert artifact_payloads(output) == artifact_payloads(FIXTURE_PATH)


def test_process_workbook_does_not_merge_comment_only_cells(tmp_path: Path) -> None:
    source = tmp_path / "comment-only.xlsx"
    output = tmp_path / "comment-only.output.xlsx"
    _write_comment_only_workbook(source)

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:A3",
        merge_empty_runs=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.filled_cells == 1
    assert result.merged_ranges == ()
    assert not output_worksheet.merged_cells.ranges
    assert output_worksheet["A1"].value == "NA"
    assert output_worksheet["A2"].comment is not None
    assert output_worksheet["A2"].comment.text == "note"


def test_process_workbook_does_not_fill_value_metadata_cells(tmp_path: Path) -> None:
    source = tmp_path / "value-metadata-only.xlsx"
    output = tmp_path / "value-metadata-only.output.xlsx"
    _write_value_metadata_only_workbook(source)

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:A3",
        merge_empty_runs=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]
    cell_attributes, cell_value = worksheet_cell_xml(output, "A2")

    assert result.filled_cells == 1
    assert result.merged_ranges == ()
    assert not output_worksheet.merged_cells.ranges
    assert output_worksheet["A1"].value == "NA"
    assert output_worksheet["A2"].value is None
    assert cell_attributes["vm"] == "1"
    assert cell_value is None
    assert output_worksheet["A3"].value == "stop"


def test_process_workbook_deletes_rows_empty_within_selected_range(tmp_path: Path) -> None:
    source = tmp_path / "delete-selected-range.xlsx"
    output = tmp_path / "delete-selected-range.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["C1"] = "outside-target"
    worksheet["A2"] = "keep"
    worksheet["A3"] = "   "
    worksheet["C3"] = "outside-target-2"
    worksheet["A4"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:B4",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 2
    assert result.filled_cells == 0
    assert result.merged_ranges == ()
    assert output_worksheet["A1"].value == "keep"
    assert output_worksheet["A2"].value == "stop"
    assert output_worksheet["C1"].value is None
    assert output_worksheet["C2"].value is None
    output_workbook.close()


def test_process_workbook_keeps_comment_rows_in_delete_mode(tmp_path: Path) -> None:
    source = tmp_path / "delete-comments.xlsx"
    output = tmp_path / "delete-comments.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A2"].comment = Comment("note", "tester")
    worksheet["A3"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:A3",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 1
    assert output_worksheet["A1"].value is None
    assert output_worksheet["A1"].comment is not None
    assert output_worksheet["A1"].comment.text == "note"
    assert output_worksheet["A2"].value == "stop"
    output_workbook.close()


def test_process_workbook_delete_mode_protects_rows_intersecting_exclusions(tmp_path: Path) -> None:
    source = tmp_path / "delete-excluded.xlsx"
    output = tmp_path / "delete-excluded.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A2"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:B2",
        excluded_ranges=["B1"],
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 0
    assert output_worksheet["A1"].value is None
    assert output_worksheet["A2"].value == "stop"
    output_workbook.close()


def test_process_workbook_delete_mode_preserves_non_empty_merged_anchor_rows(tmp_path: Path) -> None:
    source = tmp_path / "delete-merged-keep.xlsx"
    output = tmp_path / "delete-merged-keep.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "keep"
    worksheet["A3"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:B3",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 1
    assert output_worksheet["A1"].value == "keep"
    assert output_worksheet["A2"].value == "stop"
    assert worksheet_merge_refs(output) == ["A1:B1"]
    output_workbook.close()


def test_process_workbook_delete_mode_ignores_cross_row_merged_coverage(tmp_path: Path) -> None:
    source = tmp_path / "delete-merged-cross-row.xlsx"
    output = tmp_path / "delete-merged-cross-row.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.merge_cells("E1:E3")
    worksheet["E1"] = "keep-on-row-1"
    worksheet["A4"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:E4",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 2
    assert output_worksheet["E1"].value == "keep-on-row-1"
    assert output_worksheet["A2"].value == "stop"
    assert worksheet_merge_refs(output) == []
    output_workbook.close()


def test_process_workbook_delete_mode_removes_fully_empty_merged_rows(tmp_path: Path) -> None:
    source = tmp_path / "delete-merged-empty.xlsx"
    output = tmp_path / "delete-merged-empty.output.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.merge_cells("A1:B1")
    worksheet["A2"] = "stop"
    workbook.save(source)
    workbook.close()

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:B2",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert result.deleted_rows == 1
    assert output_worksheet["A1"].value == "stop"
    assert worksheet_merge_refs(output) == []
    output_workbook.close()


def test_process_workbook_delete_mode_preserves_value_metadata_cells(tmp_path: Path) -> None:
    source = tmp_path / "delete-value-metadata.xlsx"
    output = tmp_path / "delete-value-metadata.output.xlsx"
    _write_value_metadata_only_workbook(source)

    result = process_workbook(
        source,
        sheet_name="Data",
        target_range="A1:A3",
        delete_empty_rows=True,
        output_path=output,
    )

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]
    cell_attributes, cell_value = worksheet_cell_xml(output, "A1")

    assert result.deleted_rows == 1
    assert cell_attributes["vm"] == "1"
    assert cell_value is None
    assert output_worksheet["A1"].value is None
    assert output_worksheet["A2"].value == "stop"
    output_workbook.close()


def test_process_workbook_delete_mode_shifts_sheet_owned_artifacts(tmp_path: Path) -> None:
    output = tmp_path / "delete-fixture-output.xlsx"

    source_workbook = load_workbook(FIXTURE_PATH)
    source_worksheet = source_workbook["Data"]
    source_image_count = len(getattr(source_worksheet, "_images", []))
    source_chart_count = len(getattr(source_worksheet, "_charts", []))
    source_deleted_rows = build_delete_plan(
        source_worksheet,
        target_range="D2:D6",
        preserved_coordinates=find_value_metadata_cells(
            FIXTURE_PATH,
            resolve_worksheet_archive_path(FIXTURE_PATH, source_worksheet.title),
        ),
    ).deleted_row_indices
    source_hyperlink_ref = worksheet_hyperlink_refs(FIXTURE_PATH)[0]
    source_hyperlink = source_worksheet[source_hyperlink_ref].hyperlink.target
    source_workbook.close()
    source_comment_ref = comment_refs(FIXTURE_PATH)[0]
    source_threaded_comment_ref = threaded_comment_refs(FIXTURE_PATH)[0]
    source_anchor_rows = drawing_anchor_rows(FIXTURE_PATH)
    expected_hyperlink_ref = _shift_coordinate_after_deleting_rows(
        source_hyperlink_ref,
        source_deleted_rows,
    )

    result = process_workbook(
        FIXTURE_PATH,
        sheet_name="Data",
        target_range="D2:D6",
        delete_empty_rows=True,
        output_path=output,
    )

    assert result.deleted_rows == len(source_deleted_rows)
    assert result.filled_cells == 0
    assert result.merged_ranges == ()

    output_workbook = load_workbook(output)
    output_worksheet = output_workbook["Data"]

    assert output_worksheet[expected_hyperlink_ref].hyperlink is not None
    assert output_worksheet[expected_hyperlink_ref].hyperlink.target == source_hyperlink
    assert len(getattr(output_worksheet, "_images", [])) == source_image_count
    assert len(getattr(output_worksheet, "_charts", [])) == source_chart_count
    output_workbook.close()

    output_artifacts = artifact_payloads(output)
    assert set(output_artifacts) == set(artifact_payloads(FIXTURE_PATH))
    assert root_namespace_declarations(output) == root_namespace_declarations(FIXTURE_PATH)
    assert worksheet_hyperlink_refs(output) == [expected_hyperlink_ref]
    assert comment_refs(output) == [
        _shift_coordinate_after_deleting_rows(source_comment_ref, source_deleted_rows)
    ]
    assert threaded_comment_refs(output) == [
        _shift_coordinate_after_deleting_rows(source_threaded_comment_ref, source_deleted_rows)
    ]
    assert drawing_anchor_rows(output) == [
        (
            anchor_type,
            _shift_zero_based_anchor_after_deleting_rows(from_row, source_deleted_rows),
            _shift_zero_based_anchor_after_deleting_rows(to_row, source_deleted_rows),
        )
        for anchor_type, from_row, to_row in source_anchor_rows
    ]
    cell_attributes, cell_value = worksheet_cell_xml(output, "D2")
    assert cell_attributes["t"] == "e"
    assert cell_attributes["vm"] == "1"
    assert cell_value == "#VALUE!"
